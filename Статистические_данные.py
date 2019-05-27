# Import `load_workbook` module from `openpyxl`
import os
from openpyxl import load_workbook
import io
import re
from pprint import pprint
import sys
from numpy import exp, array, random, dot


pattern = re.compile(r'\w+')
PathData = ('./Data/')
ListSheet = []
PathFiles = os.listdir(path=PathData)
PathToFile = "String"
StrRead = "a"
a="String"
ValueVariable=0
n=3
Dict = {}
DictPosition = {}
ListPos = []
Gender="Gender"
Man = 'Man'
Woman = 'Woman'
Max="Max"
Min="Min"




#print(PathData)




class PathDocs(object):
    """docstring"""
 
    def __init__():
        """Constructor"""
        

    def ConfigChecker():
        with io.open('./Sys/Data.txt', encoding='utf-8') as file:
            for line in file:
                x = re.sub("^\s+|\n|\r|\s+$", '', line)
                Data = open('config.txt', 'r')
                lineS = Data.readline()
                while lineS:
                    if((len(pattern.findall(lineS)))!=0):
                        if((len(pattern.findall(lineS)))==3):
                            NameVar = str(pattern.findall(lineS)[0])
                            ValueVar = (float(str(pattern.findall(lineS)[1])))
                            Position = str(pattern.findall(lineS)[2])
                            Dict.update({NameVar: ValueVar})
                            ListPos.append(NameVar + " " + str(ValueVar) + " " + Position)
                            #DictPosition.update({NameVar:Position})

                            lineS = Data.readline()

                        if((len(pattern.findall(lineS)))==4):
                            NameVar = str(pattern.findall(lineS)[0])
                            ValueVar = (float(str(pattern.findall(lineS)[1])+"."+str(pattern.findall(lineS)[2])))
                            Position = str(pattern.findall(lineS)[3])
                            Dict.update({NameVar: ValueVar})
                            ListPos.append(NameVar + " " + str(ValueVar) + " " + Position)
                            #DictPosition.update({NameVar:Position})

                            lineS = Data.readline()
                    else:
                        break
                break

    def CheckListFiles():
        print("Начата проверка списка файлов в папке /Data/...")
        i=0
        for item in PathFiles:
            print(str(i+1)+"."+item)
            i=+1
        print("Проверка списка файлов в папке /Data/ успешно завершена...")
    def OpenFile():
        i=0
        FileName = PathFiles[int(input("Какой фаил открыть?\n"))-1]
        PathToFile = ((PathData)+str(FileName))
        print(PathToFile)
        return PathToFile
    def SheetList():
        wb = load_workbook(PathToFile)
        i=0
        #print("Найденые таблицы в документе:")
        for sheet in wb:           # go over all sheets 
            #print(str(i+1)+"."+str(sheet.title))
            ListSheet.append(sheet)
            i+=1
    def PrintData(row1,column1):
        Data = ListSheet[0].cell(row=row1, column=column1).value
        #print(Data)
        return(Data)
    def CalcDifference():
        PathDocs.PrintData(3,1)
        Data = 1
        MinCountRow=3
        MaxCountRow = 2
        MinCountColumn=1
        MaxCountColumn = 1
        kkk=0

        while str(type(Data)) != "<class 'NoneType'>":
            MaxCountRow=MaxCountRow+1
            Data = PathDocs.PrintData(MaxCountRow,1)
            if str(type(Data)) != "<class 'NoneType'>":
                #print(str(MaxCountRow) +" "+str(Data) +" " + str(type(Data)))
                MaxCountRow=MaxCountRow
            else:
                MaxCountRow=MaxCountRow-1

        Data = 1
        while str(type(Data)) != "<class 'NoneType'>":
            Data = PathDocs.PrintData(3,MaxCountColumn)
            if str(type(Data)) != "<class 'NoneType'>":
                #print(str(MaxCountColumn) +" "+str(Data) +" " + str(type(Data)))
                MaxCountColumn=MaxCountColumn
            else:
                MaxCountColumn=MaxCountColumn-1

            MaxCountColumn=MaxCountColumn+1

        print("В базе "+str(MaxCountRow)+" строк(и) и "+str(MaxCountColumn)+" столбца(ов)")
        print("Всего ячеек данных: "+str(MaxCountColumn*MaxCountRow))

        while MinCountRow != (MaxCountRow+1):
            DataList=" "
            NowGender="1"
            MinCountColumn = 1
            NowListPos=[]
            kkk=kkk+1
            #pprint(NowListPos)

            for i in range(len(ListPos)):
                if Gender in ListPos[i]:
                    #print(ListPos[i])
                    x = re.sub("^\s+|\n|\r|\s+$", '', ListPos[i])
                    PositionGender = str(pattern.findall(x)[3])
                    NowGender=PathDocs.PrintData(MinCountRow,int(PositionGender))
                    

            if(NowGender=="М"):
                for i in range(len(ListPos)):
                    if Woman in ListPos[i]:
                        continue
                    elif Gender in ListPos[i]:
                        continue
                    else:
                        NowListPos.append(ListPos[i])

            elif(NowGender=="Ж"):
                for i in range(len(ListPos)):
                    if Man in ListPos[i]:
                        continue
                    elif Gender in ListPos[i]:
                        continue
                    else:
                        NowListPos.append(ListPos[i])

            #pprint(NowListPos)
            
            while (MinCountColumn != MaxCountColumn):
                NowData = PathDocs.PrintData(int(MinCountRow),int(MinCountColumn))
                for i in range(len(NowListPos)):
                    if str(pattern.findall(NowListPos[i])[3])==str(MinCountColumn):
                        NowData=str(NowData).replace(',', '.').replace(' ', '')
                        NowData=float(NowData.strip())
                        #i1=2
                        for i1 in range(len(NowListPos)):
                            x = re.sub("^\s+|\n|\r|\s+$", '', NowListPos[i1])
                            NowData = PathDocs.PrintData(int(MinCountRow),int(MinCountColumn))
                            NowData=str(NowData)
                            NowData=NowData.replace(',', '.')
                            NowData=NowData.replace(' ', '')
                            NowData=float(NowData.strip())
                            if (str(pattern.findall(x)[3])==str(MinCountColumn)):

                                if Min in x:
                                    Calc=round((float((pattern.findall(x)[1])+"."+(pattern.findall(x)[2]))-float(NowData)),1)
                                    if (Calc>=0):
                                        print((pattern.findall(x)[0]).replace('Min', '')+" ниже нормы на "+ str(Calc)+" у № "+str(PathDocs.PrintData(int(MinCountRow),1)))
                                        #DataList=(DataList+" "+str(Calc))
                                        DataList=(DataList+" "+str("1"))
                                        MinCountColumn=MinCountColumn + 1
                                        continue
                                    elif(Calc<=0):
                                        a=pattern.findall(x)[0]
                                        #print(a)
                                        a=a.replace('Min','Max')
                                        #print(a)
                                        for i in range(len(NowListPos)):
                                            x1 = re.sub("^\s+|\n|\r|\s+$", '', NowListPos[i])
                                            if a in x1:
                                                #if Max in x:
                                                if (pattern.findall(x1)[0]==a):

                                                #print(a+" Мы считам) Сейчас:"+str(NowData)+" Стандарт: "+(pattern.findall(x)[1])+"."+(pattern.findall(x)[2]))
                                                    Calc=round((float(NowData)-float((pattern.findall(x1)[1])+"."+(pattern.findall(x1)[2]))),1)
                                                    if (Calc>=0):
                                                        print(str(pattern.findall(x1)[0])+" в таблице = "+str(NowData)+"а в конфиге = "+ (pattern.findall(x1)[1])+"."+(pattern.findall(x1)[2]))
                                                        print(pattern.findall(x1)[0].replace('Max', '')+" вSше нормы на "+ str(Calc)+" у № "+str(PathDocs.PrintData(int(MinCountRow),1)))
                                                        #DataList=(DataList+" "+str(Calc))
                                                        DataList=(DataList+" "+str(1))
                                                        MinCountColumn=MinCountColumn + 1
                                                        break
                                                    else:
                                                        print(pattern.findall(x)[0].replace('Max', '').replace('Min','')+" в норме при "+ str(NowData)+" у № "+str(PathDocs.PrintData(int(MinCountRow),1)))
                                                        DataList=(DataList+" "+str(0))

                                                        MinCountColumn=MinCountColumn + 1
                                                        break
                                        continue



                                else:
                                    Calc=round((float(NowData)-float((pattern.findall(x)[1])+"."+(pattern.findall(x)[2]))),1)
                                    if(Calc<=0):
                                        Calc=round(float((pattern.findall(x)[1])+"."+(pattern.findall(x)[2]))-float(NowData),1)
                                        DataList=(DataList+" "+str(0))
                                    else:
                                        DataList=(DataList+" "+str(1))
                                    print(pattern.findall(x)[0]+" отличается на "+ str(Calc)+" у № "+str(PathDocs.PrintData(int(MinCountRow),1)))
                                    #DataList=(DataList+" "+str(Calc))
                                    
                                    MinCountColumn=MinCountColumn + 1
                                    continue

                #print(DataList)
                MinCountColumn=MinCountColumn + 1

            fwrite = open('text.txt', 'a')
            DataList=DataList+"\n"
            fwrite.write(DataList)
            fwrite.close()
                
            
            MinCountRow=MinCountRow+1
        print(kkk)







       






                    




            
            


   


PathDocs.ConfigChecker()
PathDocs.CheckListFiles()
PathToFile = PathDocs.OpenFile()
PathDocs.SheetList()
if((input("Пройтись по таблице? 1-да 2-нет"))=="1"):
    PathDocs.CalcDifference()






class NeuralNetwork():
    def __init__(self):
        # Запустить генератор случайных чисел, чтобы он генерировал те же числа
        # каждый раз, когда программа запускается.

        random.seed(1)

        # Мы моделируем один нейрон с 3 входными и 1 выходным соединением.
        # Мы присваиваем случайные веса матрице 3 x 1 со значениями в диапазоне от -1 до 1
        # и означает 0.
        self.synaptic_weights = 2 * random.random((14, 1)) - 1

    # Сигмовидная функция, которая описывает S-образную кривую.
    # Мы передаем взвешенную сумму входных данных через эту функцию
    # нормализуйте их между 0 и 1.
    def __sigmoid(self, x):
        return 1 / (1 + exp(-x))


    # Производная сигмоидной функции.
    # Это градиент сигмовидной кривой.
    # Это показывает, насколько мы уверены в существующем весе.
    def __sigmoid_derivative(self, x):
        return x * (1 - x)


    # Мы обучаем нейронную сеть в процессе проб и ошибок.
    # Регулировка синаптических весов каждый раз.
    def train(self, training_set_inputs, training_set_outputs, number_of_training_iterations):
        for iteration in range(number_of_training_iterations):
            # Пройдите обучение через нашу нейронную сеть (один нейрон).
            output = self.think(training_set_inputs)

            # Рассчитать ошибку (разница между желаемым выводом
            # и прогнозируемый результат).
            error = training_set_outputs - output


            # Умножьте ошибку на вход и снова на градиент сигмоидной кривой.
            # Это означает, что менее уверенные веса корректируются больше.
            # Это означает, что входные данные, которые равны нулю, не вызывают изменения весов.
            adjustment = dot(training_set_inputs.T, error * self.__sigmoid_derivative(output))

            
            # Отрегулируйте вес.
            self.synaptic_weights += adjustment


    # Нейронная сеть думает.
    def think(self, inputs):
        
        # Передача входных данных через нашу нейронную сеть (наш единственный нейрон).
        return self.__sigmoid(dot(inputs, self.synaptic_weights))


if __name__ == "__main__":

    # Запуск одной нейронной сети нейронов.

    neural_network = NeuralNetwork()

    print ("Random starting synaptic weights: ")
    print (neural_network.synaptic_weights)


    # Учебный комплект. У нас есть 4 примера, каждый из которых состоит из 3 входных значений
    # и 1 выходное значение.
    #training_set_inputs = array([[0, 0, 1, 1, 1], [1, 1, 1, 1, 0], [1, 0, 1, 1, 1], [0, 1, 1, 1, 1]])
    #training_set_outputs = array([[0, 1, 1, 0]]).T

    with open('Text.txt', 'r') as file:
        lst = file.readlines()
    lst = [[int(n) for n in x.split()] for x in lst]
    lstout=[]
    for i in range(len(lst)):
        lstout.append(lst[i][-1])
        del lst[i][-1]
    print(len(array(lstout)))
    print(lstout)
    print(len(array(lst)))
    pprint(lst)

    training_set_inputs = array(lst)
    training_set_outputs = array([lstout]).T



    # Тренируйте нейронную сеть, используя тренировочный набор.
    # Сделайте это 10000 раз и вносите небольшие корректировки каждый раз.
    neural_network.train(training_set_inputs, training_set_outputs, 10000)

    print ("Новая таблица массы синапсов: ")
    print (neural_network.synaptic_weights)

    # Протестируйте нейронную сеть в новой ситуации.
    print ("Моделируем новую ситуацию [1, 0, 0] -> ?: ")
    print (neural_network.think(array([0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])))










